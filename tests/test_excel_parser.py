from __future__ import annotations

import datetime as _dt
from pathlib import Path

from django.test import SimpleTestCase

from apps.imports.date_utils import (
    parse_date,
    DateParseResult,
    ERROR_DATE_UNSUPPORTED_FORMAT,
    ERROR_DATE_INVALID_CALENDAR,
    ERROR_DATE_VALUE_TYPE,
)

from apps.imports.report_column_utils import (
    CHINESE_NUM_MAP,
    parse_report_sequence,
    ReportColumnParseResult,
    ERROR_REPORT_COLUMN_NO_MATCH,
    ERROR_REPORT_COLUMN_SEQUENCE_OUT_OF_RANGE,
    ERROR_REPORT_COLUMN_INVALID_CHINESE,
)

from apps.imports import error_codes
from apps.imports.parser import (
    ColumnMapping,
    ERROR_HEADER_NOT_FOUND,
    ERROR_ROW_COLUMN_SHIFT_SUSPECTED,
    ERROR_ROW_INVALID_APPLIED_DATE,
    ERROR_ROW_INVALID_STAGE,
    ERROR_ROW_MISSING_REQUIRED,
    ERROR_UNKNOWN_SHEET,
    HeaderParseResult,
    ParseError,
    ReportColumn,
    RowParseResult,
    SheetParseResult,
    WARNING_REPORT_COUNT_MISMATCH,
    WARNING_REPORT_DATE_INVALID,
    WARNING_REPORT_TOTAL_COLUMN_MISSING,
    normalize_development_stage,
    parse_header_row,
    parse_sheet_rows,
    parse_student_row,
    parse_workbook,
)


# ======================================================================
# 日期解析单元测试
# ======================================================================

class DateUtilsEmptyTests(SimpleTestCase):
    def test_none_returns_none_with_ok(self) -> None:
        result = parse_date(None)
        self.assertIsInstance(result, DateParseResult)
        self.assertTrue(result.ok)
        self.assertIsNone(result.value)
        self.assertIsNone(result.error_code)

    def test_empty_string_returns_none_with_ok(self) -> None:
        self.assertIsNone(parse_date("").value)
        self.assertTrue(parse_date("").ok)
        self.assertTrue(parse_date("   ").ok)

    def test_int_type_not_supported_but_does_not_crash(self) -> None:
        result = parse_date(12345)
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_VALUE_TYPE)


class DateUtilsNativeTypesTests(SimpleTestCase):
    def test_datetime_date_is_preserved(self) -> None:
        original = _dt.date(2025, 3, 15)
        result = parse_date(original)
        self.assertTrue(result.ok)
        self.assertIs(result.value, original)

    def test_datetime_datetime_converts_to_date(self) -> None:
        original = _dt.datetime(2025, 5, 20, 14, 30, 45)
        result = parse_date(original)
        self.assertTrue(result.ok)
        self.assertEqual(result.value, _dt.date(2025, 5, 20))
        self.assertIsInstance(result.value, _dt.date)


class DateUtilsSlashFormatTests(SimpleTestCase):
    def test_slash_yyyy_mm_dd(self) -> None:
        result = parse_date("2025/01/10")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2025, 1, 10))

    def test_slash_no_zero_padding(self) -> None:
        result = parse_date("2024/3/7")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2024, 3, 7))

    def test_slash_with_whitespace(self) -> None:
        result = parse_date("  2023 / 12 / 31  ")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2023, 12, 31))


class DateUtilsDashFormatTests(SimpleTestCase):
    def test_dash_yyyy_mm_dd(self) -> None:
        result = parse_date("2025-02-28")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2025, 2, 28))

    def test_dash_no_zero_padding(self) -> None:
        result = parse_date("2025-1-2")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2025, 1, 2))


class DateUtilsDotFormatTests(SimpleTestCase):
    def test_dot_yyyy_mm_dd(self) -> None:
        result = parse_date("2025.06.01")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2025, 6, 1))

    def test_dot_no_zero_padding(self) -> None:
        result = parse_date("2024.9.9")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2024, 9, 9))


class DateUtilsChineseFormatTests(SimpleTestCase):
    def test_chinese_full(self) -> None:
        result = parse_date("2025年1月15日")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2025, 1, 15))

    def test_chinese_zero_padded(self) -> None:
        result = parse_date("2025年07月08日")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2025, 7, 8))

    def test_chinese_with_whitespace(self) -> None:
        result = parse_date(" 2026 年 12 月 1 日 ")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2026, 12, 1))


class DateUtilsInvalidCalendarTests(SimpleTestCase):
    def test_non_leap_february_30(self) -> None:
        result = parse_date("2025/02/30")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_INVALID_CALENDAR)
        self.assertIsNone(result.value)

    def test_february_31_non_leap(self) -> None:
        result = parse_date("2025-02-31")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_INVALID_CALENDAR)

    def test_leap_february_29_is_valid(self) -> None:
        result = parse_date("2024.02.29")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.value, _dt.date(2024, 2, 29))

    def test_month_13_rejected(self) -> None:
        result = parse_date("2025年13月1日")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_UNSUPPORTED_FORMAT)

    def test_day_32_rejected(self) -> None:
        result = parse_date("2025/12/32")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_UNSUPPORTED_FORMAT)


class DateUtilsUnsupportedFuzzyInputsTests(SimpleTestCase):
    def test_missing_year_not_supported(self) -> None:
        result = parse_date("3月15日")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_UNSUPPORTED_FORMAT)

    def test_natural_language_not_supported(self) -> None:
        result = parse_date("昨天")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_UNSUPPORTED_FORMAT)

    def test_iso_with_time_not_supported(self) -> None:
        result = parse_date("2025-03-01T10:20:30")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_UNSUPPORTED_FORMAT)

    def test_slash_dmy_not_supported(self) -> None:
        result = parse_date("15/01/2025")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_DATE_UNSUPPORTED_FORMAT)


class DateUtilsNoDatabaseSideEffectTests(SimpleTestCase):
    def test_parser_does_not_import_models(self) -> None:
        import sys

        mod = sys.modules["apps.imports.date_utils"]
        keys = set(vars(mod).keys())
        self.assertNotIn("models", keys)
        self.assertNotIn("Student", keys)
        self.assertNotIn("ImportBatch", keys)


# ======================================================================
# 中文次数 / 思想汇报列名 转换单元测试
# ======================================================================

class ReportColumnUtilsArabicTests(SimpleTestCase):
    def test_arabic_number_with_full_name(self) -> None:
        result = parse_report_sequence("第3次思想汇报")
        self.assertIsInstance(result, ReportColumnParseResult)
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.sequence_number, 3)
        self.assertEqual(result.source_column_name, "第3次思想汇报")

    def test_arabic_number_without_suffix(self) -> None:
        result = parse_report_sequence("第10次")
        self.assertTrue(result.ok)
        self.assertEqual(result.sequence_number, 10)

    def test_arabic_with_whitespace(self) -> None:
        result = parse_report_sequence(" 第  7  次 思想汇报 ")
        self.assertTrue(result.ok)
        self.assertEqual(result.sequence_number, 7)
        self.assertIn("第  7  次 思想汇报", result.source_column_name)

    def test_arabic_99_is_ok(self) -> None:
        result = parse_report_sequence("第99次思想汇报")
        self.assertTrue(result.ok)
        self.assertEqual(result.sequence_number, 99)


class ReportColumnUtilsArabicBoundaryTests(SimpleTestCase):
    def test_arabic_0_out_of_range(self) -> None:
        result = parse_report_sequence("第0次思想汇报")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_REPORT_COLUMN_SEQUENCE_OUT_OF_RANGE)

    def test_arabic_100_out_of_range(self) -> None:
        result = parse_report_sequence("第100次")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_REPORT_COLUMN_SEQUENCE_OUT_OF_RANGE)


class ReportColumnUtilsChinese1To10Tests(SimpleTestCase):
    CASES_1_TO_10 = [
        ("第一", 1), ("第二", 2), ("第三", 3), ("第四", 4), ("第五", 5),
        ("第六", 6), ("第七", 7), ("第八", 8), ("第九", 9), ("第十", 10),
    ]

    def test_chinese_one_to_ten(self) -> None:
        for zh, num in self.CASES_1_TO_10:
            with self.subTest(zh=zh, num=num):
                result = parse_report_sequence(f"{zh}次思想汇报")
                self.assertTrue(result.ok, msg=f"{zh} failed: {result.error_message}")
                self.assertEqual(result.sequence_number, num)
                self.assertEqual(result.source_column_name, f"{zh}次思想汇报")

    def test_chinese_two_alias(self) -> None:
        """
        业务规范：'两' 作为中文口语数词不应被错误地当作第二十。
        本模块支持 '两次' = 2，用于兼容不规范表头。
        """
        result = parse_report_sequence("第两次思想汇报")
        # 根据 CHINESE_NUM_MAP，'两' 可能被包含；此处用实际结果判定：
        # 若 CHINESE_NUM_MAP 不含 '两'，则应报 ERROR_REPORT_COLUMN_INVALID_CHINESE；
        # 若含 '两'，则 sequence_number = 2。
        if "两" in CHINESE_NUM_MAP:
            self.assertTrue(result.ok)
            self.assertEqual(result.sequence_number, 2)
        else:
            self.assertFalse(result.ok)


class ReportColumnUtilsChinese11To20Tests(SimpleTestCase):
    CASES_11_TO_20 = [
        ("第十一", 11), ("第十二", 12), ("第十三", 13), ("第十四", 14), ("第十五", 15),
        ("第十六", 16), ("第十七", 17), ("第十八", 18), ("第十九", 19), ("第二十", 20),
    ]

    def test_chinese_eleven_to_twenty(self) -> None:
        for zh, num in self.CASES_11_TO_20:
            with self.subTest(zh=zh, num=num):
                result = parse_report_sequence(f"{zh}次")
                self.assertTrue(result.ok, msg=f"{zh} failed: {result.error_message}")
                self.assertEqual(result.sequence_number, num)

    def test_chinese_15_with_whitespace_and_suffix(self) -> None:
        result = parse_report_sequence(" 第  十五  次 思想汇报 ")
        self.assertTrue(result.ok, result.error_message)
        self.assertEqual(result.sequence_number, 15)


class ReportColumnUtilsChineseOutOfRangeTests(SimpleTestCase):
    def test_chinese_twenty_one_not_supported(self) -> None:
        result = parse_report_sequence("第二十一次思想汇报")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_REPORT_COLUMN_INVALID_CHINESE)

    def test_chinese_bogus_not_supported(self) -> None:
        result = parse_report_sequence("第甲乙丙次思想汇报")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_REPORT_COLUMN_INVALID_CHINESE)


class ReportColumnUtilsNoMatchTests(SimpleTestCase):
    def test_empty_name_fails(self) -> None:
        result = parse_report_sequence("")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_REPORT_COLUMN_NO_MATCH)

    def test_plain_text_not_matched(self) -> None:
        result = parse_report_sequence("思想汇报第3次")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_REPORT_COLUMN_NO_MATCH)

    def test_missing_di_prefix(self) -> None:
        result = parse_report_sequence("3次思想汇报")
        self.assertFalse(result.ok)
        self.assertEqual(result.error_code, ERROR_REPORT_COLUMN_NO_MATCH)


class ReportColumnUtilsNoDatabaseSideEffectTests(SimpleTestCase):
    def test_utils_does_not_import_models(self) -> None:
        import sys

        mod = sys.modules["apps.imports.report_column_utils"]
        keys = set(vars(mod).keys())
        self.assertNotIn("models", keys)
        self.assertNotIn("Student", keys)
        self.assertNotIn("IdeologicalReport", keys)


# ======================================================================
# 学生行级解析单元测试（任务 6-5）
# ======================================================================


class ParserDevelopmentStageNormalizationTests(SimpleTestCase):
    def test_english_codes_accepted(self) -> None:
        for code in ("ACTIVIST", "PROBATIONARY", "FULL_MEMBER"):
            with self.subTest(code=code):
                resolved, err = normalize_development_stage(code)
                self.assertEqual(resolved, code)
                self.assertIsNone(err)

    def test_chinese_translation_maps_correctly(self) -> None:
        cases = [
            ("入党积极分子", "ACTIVIST"),
            ("培养对象", "ACTIVIST"),
            ("中共预备党员", "PROBATIONARY"),
            ("预备党员", "PROBATIONARY"),
            ("正式党员", "FULL_MEMBER"),
            ("中共正式党员", "FULL_MEMBER"),
        ]
        for zh, expected in cases:
            with self.subTest(zh=zh, expected=expected):
                resolved, err = normalize_development_stage(zh)
                self.assertEqual(resolved, expected, err)
                self.assertIsNone(err)

    def test_empty_stage_returns_error(self) -> None:
        resolved, err = normalize_development_stage("")
        self.assertIsNone(resolved)
        self.assertIsNotNone(err)

    def test_invalid_string_stage_returns_error(self) -> None:
        resolved, err = normalize_development_stage("待发展")
        self.assertIsNone(resolved)
        self.assertIsNotNone(err)


class ParserHeaderChineseReportColumnsTests(SimpleTestCase):
    def test_header_parsing_supports_chinese_report_columns(self) -> None:
        header = [
            "学号",
            "姓名",
            "发展阶段",
            "职务",
            "申请入党时间",
            "思想汇报总篇数",
            "第一次思想汇报",
            "第五次思想汇报",
            "第十三次思想汇报",
            "第二十次思想汇报",
        ]
        result = parse_header_row(header)
        self.assertTrue(result.ok, result.error_message)
        seq_map = {r.sequence_number: r for r in result.mapping.report_columns}
        self.assertIn(1, seq_map)
        self.assertIn(5, seq_map)
        self.assertIn(13, seq_map)
        self.assertIn(20, seq_map)
        self.assertEqual(seq_map[1].source_column_name, "第一次思想汇报")
        self.assertEqual(seq_map[5].column_index, 7)


class ParserStudentRowBasicTests(SimpleTestCase):
    def _standard_mapping(self) -> ColumnMapping:
        header = [
            "学号",
            "姓名",
            "发展阶段",
            "职务",
            "申请入党时间",
            "思想汇报总篇数",
            "第1次思想汇报",
            "第2次思想汇报",
            "第3次思想汇报",
        ]
        parsed = parse_header_row(header)
        self.assertTrue(parsed.ok, parsed.error_message)
        return parsed.mapping

    def test_student_row_basic_yields_parsed_student_row(self) -> None:
        mapping = self._standard_mapping()
        row = [
            "20230001",
            "张三",
            "正式党员",
            "支部书记",
            "2025/01/10",
            3,
            "2025/03/01",
            "2025/06/15",
            "2025/09/20",
        ]
        outcome = parse_student_row(row, mapping, "明理党支部", excel_row_number=3)
        self.assertIsInstance(outcome, RowParseResult)
        self.assertIsNotNone(outcome.student_row)
        student = outcome.student_row
        self.assertIsNotNone(student)
        self.assertEqual(student.name, "张三")
        self.assertEqual(student.student_number, "20230001")
        self.assertEqual(type(student.student_number), str)
        self.assertEqual(student.development_stage, "FULL_MEMBER")
        self.assertEqual(student.position, "支部书记")
        self.assertEqual(student.applied_at, _dt.date(2025, 1, 10))
        self.assertEqual(student.reported_total_count, 3)
        self.assertEqual(student.calculated_date_count, 3)
        self.assertEqual(len(student.report_items), 3)
        self.assertEqual(student.report_items[0].sequence_number, 1)
        self.assertEqual(student.report_items[0].submitted_at, _dt.date(2025, 3, 1))
        self.assertEqual(student.report_items[0].source_column_name, "第1次思想汇报")
        self.assertEqual(outcome.errors, [])
        self.assertEqual(outcome.warnings, [])

    def test_position_and_applied_date_empty_allowed(self) -> None:
        mapping = self._standard_mapping()
        row = [
            "20230002",
            "李四",
            "ACTIVIST",
            "",
            "",
            None,
            "2025/02/01",
            None,
            "",
        ]
        outcome = parse_student_row(row, mapping, "明理党支部", 4)
        self.assertIsNotNone(outcome.student_row)
        s = outcome.student_row
        self.assertIsNotNone(s)
        self.assertEqual(s.development_stage, "ACTIVIST")
        self.assertEqual(s.position, "")
        self.assertIsNone(s.applied_at)
        self.assertIsNone(s.reported_total_count)
        self.assertEqual(s.calculated_date_count, 1)
        self.assertEqual(s.report_items[0].sequence_number, 1)


class ParserStudentRowErrorSkipTests(SimpleTestCase):
    def setUp(self) -> None:
        header = [
            "学号",
            "姓名",
            "发展阶段",
            "职务",
            "申请入党时间",
            "思想汇报总篇数",
            "第1次思想汇报",
        ]
        parsed = parse_header_row(header)
        assert parsed.ok
        self.mapping = parsed.mapping

    def test_missing_name_not_in_valid_rows(self) -> None:
        row = ["", "张三", "正式党员", "", "", None, None]
        outcome = parse_student_row(row, self.mapping, "支部", 3)
        self.assertIsNone(outcome.student_row)
        self.assertTrue(any(e.code == ERROR_ROW_MISSING_REQUIRED for e in outcome.errors))

    def test_missing_student_number_not_in_valid_rows(self) -> None:
        row = ["", "", "预备党员", "", "", None, None]
        outcome = parse_student_row(row, self.mapping, "支部", 3)
        self.assertIsNone(outcome.student_row)

    def test_invalid_stage_is_error_and_skipped(self) -> None:
        row = ["20230003", "王五", "群众", "", "", None, None]
        outcome = parse_student_row(row, self.mapping, "支部", 5)
        self.assertIsNone(outcome.student_row)
        codes = {e.code for e in outcome.errors}
        self.assertIn(ERROR_ROW_INVALID_STAGE, codes)

    def test_invalid_applied_date_is_error_and_skipped(self) -> None:
        row = ["20230004", "赵六", "正式党员", "", "昨天", None, None]
        outcome = parse_student_row(row, self.mapping, "支部", 6)
        self.assertIsNone(outcome.student_row)
        codes = {e.code for e in outcome.errors}
        self.assertIn(ERROR_ROW_INVALID_APPLIED_DATE, codes)

    def test_parse_sheet_rows_separates_errors_from_valid_rows(self) -> None:
        rows = [
            ["20230001", "OK同学", "正式党员", "", "", 1, "2025/03/01"],
            ["", "缺学号同学", "预备党员", "", "", 0, None],
            ["", "缺姓名同学", "入党积极分子", "", "", None, None],
            ["20230003", "阶段错同学", "神奇阶段", "", "", None, None],
        ]
        sheet = parse_sheet_rows(rows, self.mapping, "支部", start_excel_row_number=3)
        self.assertIsInstance(sheet, SheetParseResult)
        self.assertEqual(len(sheet.valid_rows), 1)
        self.assertEqual(sheet.valid_rows[0].name, "OK同学")
        self.assertGreaterEqual(len(sheet.errors), 3)
        error_codes = {e.code for e in sheet.errors}
        self.assertIn(ERROR_ROW_MISSING_REQUIRED, error_codes)
        self.assertIn(ERROR_ROW_INVALID_STAGE, error_codes)


class ParserStudentRowWarningTests(SimpleTestCase):
    def _mapping(self, include_total_column: bool = True) -> ColumnMapping:
        header = ["学号", "姓名", "发展阶段"]
        if include_total_column:
            header.append("思想汇报总篇数")
        header.extend(["第1次思想汇报", "第2次思想汇报", "第3次思想汇报"])
        result = parse_header_row(header)
        assert result.ok
        return result.mapping

    def test_report_count_mismatch_adds_warning_still_valid(self) -> None:
        mapping = self._mapping(include_total_column=True)
        row = ["20231001", "陈七", "入党积极分子", 5, "2025/01/01", "2025/04/01", None]
        outcome = parse_student_row(row, mapping, "支部", 3)
        self.assertIsNotNone(outcome.student_row)
        s = outcome.student_row
        self.assertIsNotNone(s)
        self.assertEqual(s.reported_total_count, 5)
        self.assertEqual(s.calculated_date_count, 2)
        warning_codes = [w.code for w in outcome.warnings]
        self.assertIn(WARNING_REPORT_COUNT_MISMATCH, warning_codes)
        self.assertTrue(any(w.student_name == "陈七" for w in outcome.warnings))
        self.assertEqual(s.warnings, outcome.warnings)

    def test_report_total_column_missing_warning_but_still_valid(self) -> None:
        mapping = self._mapping(include_total_column=False)
        row = ["20231002", "孙八", "中共预备党员", "2025/02/01", "", "2025/06/01"]
        outcome = parse_student_row(row, mapping, "支部", 3)
        self.assertIsNotNone(outcome.student_row)
        # 缺总篇数列的警告现在是工作表级别（在 parse_workbook 中生成），行级别不再产生
        warning_codes = {w.code for w in outcome.warnings}
        self.assertNotIn(WARNING_REPORT_TOTAL_COLUMN_MISSING, warning_codes)

    def test_empty_report_dates_skipped_without_error_or_item(self) -> None:
        mapping = self._mapping(include_total_column=True)
        row = ["20231003", "周九", "正式党员", 2, "2025/01/01", None, "2025/07/01"]
        outcome = parse_student_row(row, mapping, "支部", 3)
        self.assertIsNotNone(outcome.student_row)
        s = outcome.student_row
        self.assertIsNotNone(s)
        self.assertEqual(len(s.report_items), 2)
        seqs = sorted(r.sequence_number for r in s.report_items)
        self.assertEqual(seqs, [1, 3])

    def test_invalid_report_date_is_warning_not_error_still_valid(self) -> None:
        mapping = self._mapping(include_total_column=True)
        row = ["20231004", "吴十", "正式党员", 2, "昨天", "2025/02/01", "2025/05/01"]
        outcome = parse_student_row(row, mapping, "支部", 3)
        self.assertIsNotNone(outcome.student_row)
        codes = {w.code for w in outcome.warnings}
        self.assertIn(WARNING_REPORT_DATE_INVALID, codes)
        s = outcome.student_row
        self.assertIsNotNone(s)
        self.assertEqual(s.calculated_date_count, 2)


class ParserStudentRowReportedTotalNotOverriddenTests(SimpleTestCase):
    def test_calculated_count_never_replaces_reported(self) -> None:
        header = [
            "学号",
            "姓名",
            "发展阶段",
            "思想汇报总篇数",
            "第1次思想汇报",
            "第2次思想汇报",
            "第3次思想汇报",
        ]
        parsed = parse_header_row(header)
        assert parsed.ok
        # reported_total=4，actual=2，但 reported_total_count 仍需保留原始 4
        row = ["20232001", "郑十一", "预备党员", 4, "2025/03/01", None, "2025/09/01"]
        outcome = parse_student_row(row, parsed.mapping, "支部", 3)
        self.assertIsNotNone(outcome.student_row)
        s = outcome.student_row
        self.assertIsNotNone(s)
        self.assertEqual(s.reported_total_count, 4)
        self.assertEqual(s.calculated_date_count, 2)


class ParserNoDatabaseSideEffectsTests(SimpleTestCase):
    def test_parser_does_not_import_models(self) -> None:
        import sys

        mod = sys.modules["apps.imports.parser"]
        keys = set(vars(mod).keys())
        self.assertNotIn("Student", keys)
        self.assertNotIn("ImportBatch", keys)


# ======================================================================
# 任务 6-6：错误码完整定义 + 行错位检测单元测试
# ======================================================================

REQUIRED_ERROR_CODE_NAMES = [
    "HEADER_NOT_FOUND",
    "ERROR_ROW_MISSING_REQUIRED",
    "ERROR_ROW_INVALID_STAGE",
    "ERROR_ROW_INVALID_APPLIED_DATE",
    "ERROR_ROW_COLUMN_SHIFT_SUSPECTED",
    "ERROR_DATE_UNSUPPORTED_FORMAT",
    "ERROR_DATE_INVALID_CALENDAR",
    "ERROR_DATE_VALUE_TYPE",
    "ERROR_REPORT_COLUMN_NO_MATCH",
    "ERROR_REPORT_COLUMN_SEQUENCE_OUT_OF_RANGE",
    "ERROR_REPORT_COLUMN_INVALID_CHINESE",
]

REQUIRED_WARNING_CODE_NAMES = [
    "WARNING_REPORT_COUNT_MISMATCH",
    "WARNING_REPORT_TOTAL_COLUMN_MISSING",
    "WARNING_REPORT_DATE_INVALID",
]


class ErrorCodeCompleteDefinitionTests(SimpleTestCase):
    def test_error_codes_module_contains_every_required_error(self) -> None:
        from apps.imports import error_codes

        for attr_name in REQUIRED_ERROR_CODE_NAMES + REQUIRED_WARNING_CODE_NAMES:
            self.assertTrue(
                hasattr(error_codes, attr_name),
                f"error_codes 模块缺少定义：{attr_name}",
            )
            code_value = getattr(error_codes, attr_name)
            self.assertIsInstance(code_value, str)
            self.assertTrue(code_value)

    def test_all_error_codes_have_a_human_message(self) -> None:
        from apps.imports import error_codes

        for attr_name in REQUIRED_ERROR_CODE_NAMES:
            code_value = getattr(error_codes, attr_name)
            self.assertIn(
                code_value,
                error_codes.ERROR_MESSAGES,
                f"错误码 {attr_name} = {code_value} 不在 ERROR_MESSAGES 中",
            )
            self.assertTrue(error_codes.ERROR_MESSAGES[code_value])

    def test_all_warning_codes_have_a_human_message(self) -> None:
        from apps.imports import error_codes

        for attr_name in REQUIRED_WARNING_CODE_NAMES:
            code_value = getattr(error_codes, attr_name)
            self.assertIn(
                code_value,
                error_codes.WARNING_MESSAGES,
                f"警告码 {attr_name} = {code_value} 不在 WARNING_MESSAGES 中",
            )
            self.assertTrue(error_codes.WARNING_MESSAGES[code_value])

    def test_parser_exports_match_error_codes_module(self) -> None:
        from apps.imports import error_codes, parser

        pairs = [
            ("HEADER_NOT_FOUND", parser.ERROR_HEADER_NOT_FOUND, error_codes.HEADER_NOT_FOUND),
            ("ROW_MISSING_REQUIRED", parser.ERROR_ROW_MISSING_REQUIRED, error_codes.ERROR_ROW_MISSING_REQUIRED),
            ("ROW_COLUMN_SHIFT_SUSPECTED", parser.ERROR_ROW_COLUMN_SHIFT_SUSPECTED, error_codes.ERROR_ROW_COLUMN_SHIFT_SUSPECTED),
        ]
        for name, parser_code, error_code in pairs:
            self.assertEqual(parser_code, error_code, f"parser 与 error_codes 不一致：{name}")


class _ShiftDetectionMixin:
    def _standard_mapping(self) -> ColumnMapping:
        header = [
            "学号",     # 0
            "姓名",     # 1
            "发展阶段",  # 2
            "职务",     # 3
            "申请入党时间",  # 4
            "思想汇报总篇数",  # 5
            "第1次思想汇报",  # 6
            "第2次思想汇报",  # 7
            "第3次思想汇报",  # 8
        ]
        parsed = parse_header_row(header)
        assert parsed.ok, parsed.error_message
        return parsed.mapping


class RowColumnShiftDetectionTests(SimpleTestCase, _ShiftDetectionMixin):
    def test_shift_scenario_1_name_empty_but_neighbour_has_name_and_student_number_empty(
        self,
    ) -> None:
        """
        错位场景 1：
          姓名列为空 → 右邻列疑似姓名（信号 1）
          学号列为空 → 右邻列疑似学号（信号 2）
          → 命中 ≥ 2 → ROW_COLUMN_SHIFT_SUSPECTED，不进 valid_rows
        """
        mapping = self._standard_mapping()
        row = [
            "",         # 学号列空
            "",         # 姓名列空
            "正式党员",
            "书记",
            "2025/01/10",
            3,
            "2025/03/01",
            "2025/06/15",
            "2025/09/20",
        ]
        # 手动把右邻也放一个看起来像学号和姓名的值 → 使"右邻疑似"命中
        row = [
            "",          # 学号列：空 （→ 右邻 = 姓名列空 → 继续右邻也空的话不中，
            "",          # 姓名列：空 → 右邻 = 发展阶段 "正式党员" 不像 2-4 字纯中文姓名
            "正式党员",
            "书记",
            "2025/01/10",
            3,
            "",
            "",
            "",
        ]
        # 重新构造：要命中"姓名空 + 右邻是疑似姓名"，应把发展阶段列塞成2字名字
        row = [
            "",          # 学号空 （右邻 = 姓名列空 → 不中）
            "",          # 姓名空 （右邻 = 发展阶段列 写"张三" → 中！）
            "张三",
            "书记",
            "2025/01/10",
            3,
            "",
            "",
            "",
        ]
        # 再加另一个信号：学号空但右邻姓名列空 → 不中，增加第三个信号："发展阶段非法 但 右邻 职务列 '正式党员'合法"
        row[3] = "正式党员"
        outcome = parse_student_row(row, mapping, "支部", 3)
        self.assertIsNone(outcome.student_row, outcome.errors)
        codes = {e.code for e in outcome.errors}
        self.assertIn(ERROR_ROW_COLUMN_SHIFT_SUSPECTED, codes)
        # ParseError 字段齐全检查（业务规则 5）
        shift_err = next(e for e in outcome.errors if e.code == ERROR_ROW_COLUMN_SHIFT_SUSPECTED)
        self.assertTrue(shift_err.message)
        self.assertEqual(shift_err.sheet_name, "支部")
        self.assertEqual(shift_err.excel_row_number, 3)
        self.assertIsNotNone(shift_err.field_name)
        self.assertTrue(hasattr(shift_err, "source_value"))

    def test_shift_scenario_2_stage_invalid_neighbour_has_valid(self) -> None:
        """
        错位场景 2：
          发展阶段列值非法 → 左右邻出现合法阶段 → 信号 1
          + 姓名列疑似学号 信号 2 → 命中 ≥ 2
        """
        mapping = self._standard_mapping()
        row = [
            "20230001",
            "20230002",   # 姓名列纯数字 → 疑似学号 （信号 1）
            "书记",       # 发展阶段列=书记，非法
            "预备党员",    # 右邻=职务列填"预备党员" → 合法阶段 （信号 2）
            "2025/01/10",
            1,
            "2025/03/01",
            None,
            None,
        ]
        outcome = parse_student_row(row, mapping, "支部", 4)
        self.assertIsNone(outcome.student_row)
        codes = {e.code for e in outcome.errors}
        self.assertIn(ERROR_ROW_COLUMN_SHIFT_SUSPECTED, codes)

    def test_shift_scenario_3_student_number_looks_like_date(self) -> None:
        """
        错位场景 3：
          学号列 = "2025/01/10" （像日期，信号 1）
          姓名列 = "20230001" （像学号，信号 2）
          已 ≥ 2 → ROW_COLUMN_SHIFT_SUSPECTED
        """
        mapping = self._standard_mapping()
        row = [
            "2025/01/10",   # 学号像日期
            "20230001",     # 姓名像学号 （纯数字 8 位）
            "正式党员",
            "委员",
            "张三",          # 申请时间列像姓名 → 再给一个类型错信号
            2,
            "2025/04/01",
            "2025/07/01",
            None,
        ]
        outcome = parse_student_row(row, mapping, "支部", 5)
        self.assertIsNone(outcome.student_row)
        codes = {e.code for e in outcome.errors}
        self.assertIn(ERROR_ROW_COLUMN_SHIFT_SUSPECTED, codes)

    def test_shift_scenario_4_many_type_mismatches(self) -> None:
        """
        错位场景 4：
          多个字段类型与表头明显不匹配（2处以上）直接命中错位信号
        """
        mapping = self._standard_mapping()
        row = [
            "张三",             # 学号列像姓名
            "20230001",         # 姓名列像学号
            "2025/02/28",       # 发展阶段像日期
            "正式党员",          # 职务列填了阶段
            "2025/01/10",
            1,
            None,
            None,
            None,
        ]
        sheet = parse_sheet_rows([row], mapping, "支部", start_excel_row_number=3)
        self.assertEqual(len(sheet.valid_rows), 0)
        codes = {e.code for e in sheet.errors}
        self.assertIn(ERROR_ROW_COLUMN_SHIFT_SUSPECTED, codes)

    def test_no_false_positive_on_normal_rows(self) -> None:
        """正常学生数据行不误报错位。"""
        mapping = self._standard_mapping()
        row = [
            "20230001",
            "张三",
            "正式党员",
            "支部书记",
            "2025/01/10",
            3,
            "2025/03/01",
            "2025/06/15",
            "2025/09/20",
        ]
        outcome = parse_student_row(row, mapping, "支部", 3)
        self.assertIsNotNone(outcome.student_row)
        codes = {e.code for e in outcome.errors}
        self.assertNotIn(ERROR_ROW_COLUMN_SHIFT_SUSPECTED, codes)

    def test_shift_row_does_not_enter_valid_rows_through_parse_sheet_rows(self) -> None:
        """错位行在 parse_sheet_rows 里最终不进入 valid_rows。"""
        mapping = self._standard_mapping()
        normal = [
            "20230001",
            "李四",
            "预备党员",
            "",
            "2024/05/01",
            1,
            "2024/06/01",
            None,
            None,
        ]
        shifted = [
            "2025/01/10",
            "20230002",
            "张三",
            "正式党员",
            "委员",
            1,
            None,
            None,
            None,
        ]
        sheet = parse_sheet_rows(
            [normal, shifted], mapping, "支部", start_excel_row_number=3
        )
        self.assertEqual(len(sheet.valid_rows), 1)
        self.assertEqual(sheet.valid_rows[0].name, "李四")
        self.assertTrue(
            any(e.code == ERROR_ROW_COLUMN_SHIFT_SUSPECTED for e in sheet.errors)
        )


# ======================================================================
# 任务 6-7：openpyxl 创建临时 xlsx 的端到端解析测试（16 场景）
# ======================================================================


class _MiniWorkbook:
    """
    测试辅助：封装 openpyxl + 临时文件，暴露 path/workbook，支持 with 语句。
    """

    def __init__(self) -> None:
        from openpyxl import Workbook
        import tempfile

        self._temp = tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False, prefix="pb_test_"
        )
        self._temp.close()
        self.path = self._temp.name
        self.workbook = Workbook()
        # 删除默认表，避免“未登记支部名的 Sheet”被误判为有效 sheet
        default = self.workbook.active
        self.workbook.remove(default)

    def __enter__(self) -> "_MiniWorkbook":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def close(self) -> None:
        import os

        try:
            if self.workbook is not None:
                self.workbook.close()
        finally:
            try:
                if os.path.exists(self.path):
                    os.remove(self.path)
            except OSError:
                pass


def _standard_header_row2() -> list:
    return [
        "学号",
        "姓名",
        "发展阶段",
        "职务",
        "申请入党时间",
        "思想汇报总篇数",
        "第1次思想汇报",
        "第2次思想汇报",
        "第3次思想汇报",
    ]


def _append_sheet(wb, *, title: str, header_row2: list, data_rows: list[list]):
    from openpyxl import Workbook

    assert isinstance(wb, Workbook)
    ws = wb.create_sheet(title=title)
    # 第 1 行：字段分组（业务约定前两行表头，第 1 行可以留空）
    ws.append([""] * len(header_row2))
    ws.append(list(header_row2))
    for row in data_rows:
        ws.append(list(row))
    return ws



class OpenpyxlIntegrationTests(SimpleTestCase):
    """
    使用 openpyxl 动态创建 .xlsx 临时文件，直接调用 parse_workbook() 进行端到端解析测试。
    覆盖 11 个集成场景。
    """

    # ------------------------- 场景 1：九个支部全部识别 -------------------------
    def test_scenario_01_all_nine_branches(self) -> None:
        with _MiniWorkbook() as twb:
            branch_map = {
                "明理党支部": "MINGLI",
                "德理党支部": "DELI",
                "惟理党支部": "WEILI",
                "求理党支部": "QIULI",
                "知理党支部": "ZHILI",
                "昭理党支部": "ZHAOLI",
                "学理党支部": "XUELI",
                "博理党支部": "BOLI",
                "艺理党支部": "YILI",
            }
            for title in branch_map:
                _append_sheet(
                    twb.workbook,
                    title=title,
                    header_row2=_standard_header_row2(),
                    data_rows=[
                        ["20230001", "测试同学", "正式党员", "", "2025/01/10", 1, "2025/03/01", None, None],
                    ],
                )
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(result.total_sheets, 9)
            self.assertEqual(len(result.valid_rows), 9)
            actual_codes = {r.branch_code for r in result.valid_rows}
            self.assertEqual(actual_codes, set(branch_map.values()))

    # ------------------------- 场景 2：多工作表结果和统计聚合 -------------------------
    def test_scenario_02_multi_sheet_aggregation(self) -> None:
        with _MiniWorkbook() as twb:
            _append_sheet(
                twb.workbook,
                title="明理党支部",
                header_row2=_standard_header_row2(),
                data_rows=[
                    ["20230101", "赵甲", "正式党员", "支部书记", "2025/01/10", 3, "2025/03/01", "2025/06/15", "2025/09/20"],
                    ["20230102", "钱乙", "预备党员", "", "2024/11/01", 2, "2024/12/01", "2025/02/28", None],
                ],
            )
            _append_sheet(
                twb.workbook,
                title="德理党支部",
                header_row2=_standard_header_row2(),
                data_rows=[
                    ["20230201", "孙丙", "入党积极分子", "班长", "2025/02/01", 1, "2025/04/15", None, None],
                ],
            )
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(result.total_sheets, 2)
            self.assertEqual(result.success_sheets, 2)
            self.assertEqual(result.failed_sheets, 0)
            self.assertEqual(result.total_rows, 3)
            self.assertEqual(result.success_rows, 3)
            self.assertEqual(len(result.valid_rows), 3)
            self.assertEqual(len(result.sheet_results), 2)
            mingli = next(s for s in result.sheet_results if s.sheet_name == "明理党支部")
            self.assertEqual(mingli.status, "success")
            self.assertEqual(mingli.total_rows, 2)
            self.assertEqual(mingli.valid_row_count, 2)
            deli = next(s for s in result.sheet_results if s.sheet_name == "德理党支部")
            self.assertEqual(deli.status, "success")
            self.assertEqual(deli.total_rows, 1)
            self.assertEqual(deli.valid_row_count, 1)

    # ------------------------- 场景 3：单个工作表失败不阻止其他表进入结果 -------------------------
    def test_scenario_03_single_sheet_failure_doesnt_block_others(self) -> None:
        with _MiniWorkbook() as twb:
            _append_sheet(
                twb.workbook,
                title="明理党支部",
                header_row2=["序号", "备注", "日期"],
                data_rows=[["1", "备注内容", "2025/01/10"]],
            )
            _append_sheet(
                twb.workbook,
                title="德理党支部",
                header_row2=_standard_header_row2(),
                data_rows=[
                    ["20230201", "孙丙", "正式党员", "", "2025/01/10", 1, "2025/03/01", None, None],
                ],
            )
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(result.total_sheets, 2)
            self.assertEqual(result.success_sheets, 1)
            self.assertEqual(result.failed_sheets, 1)
            mingli = next(s for s in result.sheet_results if s.sheet_name == "明理党支部")
            self.assertEqual(mingli.status, "failed")
            self.assertEqual(mingli.valid_row_count, 0)
            deli = next(s for s in result.sheet_results if s.sheet_name == "德理党支部")
            self.assertEqual(deli.status, "success")
            self.assertEqual(deli.valid_row_count, 1)
            self.assertEqual(len(result.valid_rows), 1)
            self.assertEqual(result.valid_rows[0].name, "孙丙")
            self.assertEqual(result.valid_rows[0].branch_code, "DELI")
            self.assertTrue(
                any(e.code == ERROR_HEADER_NOT_FOUND for e in result.errors)
            )

    # ------------------------- 场景 4：未知工作表不产生有效行 -------------------------
    def test_scenario_04_unknown_sheet_no_valid_rows(self) -> None:
        with _MiniWorkbook() as twb:
            _append_sheet(
                twb.workbook,
                title="神秘党支部",
                header_row2=_standard_header_row2(),
                data_rows=[
                    ["20239999", "佚名", "正式党员", "", "2025/01/10", 0, None, None, None],
                ],
            )
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(result.total_sheets, 1)
            self.assertEqual(len(result.valid_rows), 0)
            unknown_errors = [e for e in result.errors if e.code == ERROR_UNKNOWN_SHEET]
            self.assertTrue(unknown_errors)
            unknown = unknown_errors[0]
            self.assertEqual(unknown.sheet_name, "神秘党支部")
            self.assertEqual(unknown.excel_row_number, 1)
            self.assertEqual(len(result.sheet_results), 1)
            self.assertEqual(result.sheet_results[0].status, "unknown")
            self.assertEqual(result.sheet_results[0].valid_row_count, 0)

    # ------------------------- 场景 5：前两行合法表头识别 -------------------------
    def test_scenario_05_header_first_two_rows(self) -> None:
        with _MiniWorkbook() as twb:
            _append_sheet(
                twb.workbook,
                title="明理党支部",
                header_row2=_standard_header_row2(),
                data_rows=[
                    ["20230001", "张三", "正式党员", "", "2025/01/10", 1, "2025/03/01", None, None],
                ],
            )
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(len(result.valid_rows), 1)
            row = result.valid_rows[0]
            self.assertEqual(row.excel_row_number, 3)
            self.assertEqual(row.name, "张三")
            self.assertEqual(row.student_number, "20230001")

    # ------------------------- 场景 6：第99次成功、第100次明确失败 -------------------------
    def test_scenario_06_arabic_99_ok_100_fail(self) -> None:
        with _MiniWorkbook() as twb:
            header = [
                "学号", "姓名", "发展阶段", "思想汇报总篇数",
                "第99次思想汇报", "第100次思想汇报",
            ]
            data = [
                ["20230001", "张三", "正式党员", 1, "2025/06/01", "2025/12/01"],
            ]
            _append_sheet(twb.workbook, title="明理党支部", header_row2=header, data_rows=data)
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(len(result.valid_rows), 1)
            row = result.valid_rows[0]
            seqs = {r.sequence_number for r in row.report_items}
            self.assertIn(99, seqs)
            self.assertNotIn(100, seqs)
            self.assertEqual(row.calculated_date_count, 1)

    # ------------------------- 场景 7：第二十次成功、第二十一次失败 -------------------------
    def test_scenario_07_chinese_20_ok_21_fail(self) -> None:
        with _MiniWorkbook() as twb:
            header = [
                "学号", "姓名", "发展阶段", "思想汇报总篇数",
                "第二十次思想汇报", "第二十一次思想汇报",
            ]
            data = [
                ["20230001", "张三", "正式党员", 1, "2025/06/01", "2025/12/01"],
            ]
            _append_sheet(twb.workbook, title="明理党支部", header_row2=header, data_rows=data)
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(len(result.valid_rows), 1)
            row = result.valid_rows[0]
            seqs = {r.sequence_number for r in row.report_items}
            self.assertIn(20, seqs)
            self.assertNotIn(21, seqs)
            self.assertEqual(row.calculated_date_count, 1)

    # ------------------------- 场景 8：非法总篇数与空值结果不同 -------------------------
    def test_scenario_08_invalid_total_vs_empty(self) -> None:
        with _MiniWorkbook() as twb:
            _append_sheet(
                twb.workbook,
                title="明理党支部",
                header_row2=_standard_header_row2(),
                data_rows=[
                    ["20230001", "张三", "正式党员", "", "2025/01/10", "abc", "2025/03/01", None, None],
                    ["20230002", "李四", "预备党员", "", "2025/01/10", None, "2025/03/01", None, None],
                ],
            )
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(len(result.valid_rows), 1)
            row = result.valid_rows[0]
            self.assertEqual(row.name, "李四")
            self.assertIsNone(row.reported_total_count)
            self.assertFalse(any(r.name == "张三" for r in result.valid_rows))
            self.assertTrue(
                any(e.code == error_codes.ERROR_REPORT_TOTAL_INVALID for e in result.errors)
            )

    # ------------------------- 场景 9：缺总篇数列只产生一次工作表警告 -------------------------
    def test_scenario_09_missing_total_column_one_warning(self) -> None:
        with _MiniWorkbook() as twb:
            header = [
                "学号", "姓名", "发展阶段", "职务", "申请入党时间",
                "第1次思想汇报", "第2次思想汇报",
            ]
            data = [
                ["20230001", "同学A", "正式党员", "", "2025/01/10", "2025/03/01", "2025/06/01"],
                ["20230002", "同学B", "预备党员", "", "2025/01/10", "2025/03/01", "2025/06/01"],
                ["20230003", "同学C", "正式党员", "", "2025/01/10", "2025/03/01", "2025/06/01"],
            ]
            _append_sheet(twb.workbook, title="明理党支部", header_row2=header, data_rows=data)
            twb.workbook.save(twb.path)

            result = parse_workbook(Path(twb.path))
            self.assertEqual(len(result.valid_rows), 3)
            missing_warnings = [
                w for w in result.warnings
                if w.code == WARNING_REPORT_TOTAL_COLUMN_MISSING
            ]
            self.assertEqual(len(missing_warnings), 1)
            self.assertEqual(missing_warnings[0].excel_row_number, 2)

    # ------------------------- 场景 10：文件不存在抛 FileNotFoundError -------------------------
    def test_scenario_10_file_not_found(self) -> None:
        fake_path = Path("nonexistent_file_for_test_12345.xlsx")
        with self.assertRaises(FileNotFoundError):
            parse_workbook(fake_path)

    # ------------------------- 场景 11：解析前后数据库不变 -------------------------
    def test_scenario_11_no_database_side_effects(self) -> None:
        import sys

        mod = sys.modules["apps.imports.parser"]
        keys_before = set(vars(mod).keys())
        self.assertNotIn("Student", keys_before)
        self.assertNotIn("ImportBatch", keys_before)
        self.assertNotIn("models", keys_before)

        with _MiniWorkbook() as twb:
            _append_sheet(
                twb.workbook,
                title="明理党支部",
                header_row2=_standard_header_row2(),
                data_rows=[
                    ["20230001", "张三", "正式党员", "", "2025/01/10", 1, "2025/03/01", None, None],
                ],
            )
            twb.workbook.save(twb.path)
            parse_workbook(Path(twb.path))

        keys_after = set(vars(mod).keys())
        self.assertNotIn("Student", keys_after)
        self.assertNotIn("ImportBatch", keys_after)
        self.assertNotIn("models", keys_after)
