from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from apps.imports import error_codes
from apps.imports.datatypes import (
    ParseError,
    ParseResult,
    ParseWarning,
    ParsedReportItem,
    ParsedStudentRow,
    SheetResult,
)
from apps.imports.date_utils import parse_date
from apps.imports.report_column_utils import parse_report_sequence

# 向后兼容：所有错误码和警告码实际来源统一为 error_codes 模块
ERROR_HEADER_NOT_FOUND = error_codes.ERROR_HEADER_NOT_FOUND
ERROR_ROW_MISSING_REQUIRED = error_codes.ERROR_ROW_MISSING_REQUIRED
ERROR_ROW_INVALID_STAGE = error_codes.ERROR_ROW_INVALID_STAGE
ERROR_ROW_INVALID_APPLIED_DATE = error_codes.ERROR_ROW_INVALID_APPLIED_DATE
ERROR_ROW_COLUMN_SHIFT_SUSPECTED = error_codes.ERROR_ROW_COLUMN_SHIFT_SUSPECTED
ERROR_DATE_UNSUPPORTED_FORMAT = error_codes.ERROR_DATE_UNSUPPORTED_FORMAT
ERROR_DATE_INVALID_CALENDAR = error_codes.ERROR_DATE_INVALID_CALENDAR
ERROR_DATE_VALUE_TYPE = error_codes.ERROR_DATE_VALUE_TYPE
ERROR_REPORT_COLUMN_NO_MATCH = error_codes.ERROR_REPORT_COLUMN_NO_MATCH
ERROR_REPORT_COLUMN_SEQUENCE_OUT_OF_RANGE = error_codes.ERROR_REPORT_COLUMN_SEQUENCE_OUT_OF_RANGE
ERROR_REPORT_COLUMN_INVALID_CHINESE = error_codes.ERROR_REPORT_COLUMN_INVALID_CHINESE

WARNING_REPORT_COUNT_MISMATCH = error_codes.WARNING_REPORT_COUNT_MISMATCH
WARNING_REPORT_TOTAL_COLUMN_MISSING = error_codes.WARNING_REPORT_TOTAL_COLUMN_MISSING
WARNING_REPORT_DATE_INVALID = error_codes.WARNING_REPORT_DATE_INVALID
ERROR_UNKNOWN_SHEET = error_codes.ERROR_UNKNOWN_SHEET

ALL_ERROR_CODES = error_codes.ALL_ERROR_CODES
ALL_WARNING_CODES = error_codes.ALL_WARNING_CODES
ERROR_MESSAGES = error_codes.ERROR_MESSAGES
WARNING_MESSAGES = error_codes.WARNING_MESSAGES

_MSG_NAME_MISSING = "无法识别表头：缺少核心字段【姓名】"
_MSG_STUDENT_NUMBER_MISSING = "无法识别表头：缺少核心字段【学号】"
_MSG_STAGE_MISSING = "无法识别表头：缺少核心字段【发展阶段】"
_MSG_ROW_NAME_EMPTY = "姓名为空"
_MSG_ROW_STUDENT_NUMBER_EMPTY = "学号为空"
_MSG_ROW_STAGE_EMPTY = "发展阶段为空"
_MSG_ROW_INVALID_STAGE = "发展阶段无效，仅允许：ACTIVIST/PROBATIONARY/FULL_MEMBER 或其中文对应值"
_MSG_ROW_INVALID_APPLIED_DATE = "申请入党时间格式错误"
_MSG_REPORT_COUNT_MISMATCH_TEMPLATE = "思想汇报总篇数填报({reported})与系统计算({calculated})不一致"
_MSG_REPORT_TOTAL_COLUMN_MISSING = "缺少【思想汇报总篇数】列，无法校验总篇数"
_MSG_REPORT_DATE_INVALID_TEMPLATE = "思想汇报【{source_column}】日期无法解析，已跳过"

NAME_ALIASES: tuple[str, ...] = ("姓名", "学生姓名", "名字")
STUDENT_NUMBER_ALIASES: tuple[str, ...] = ("学号", "学生学号", "学生编号", "学号/工号")
DEVELOPMENT_STAGE_ALIASES: tuple[str, ...] = (
    "发展阶段",
    "身份",
    "党员身份",
    "发展状态",
    "培养阶段",
)
POSITION_ALIASES: tuple[str, ...] = (
    "职务",
    "党内职务",
    "担任职务",
    "担任党内职务",
)
APPLIED_AT_ALIASES: tuple[str, ...] = (
    "申请入党时间",
    "入党申请时间",
    "申请入党日期",
    "递交申请书时间",
)
REPORTED_TOTAL_COUNT_ALIASES: tuple[str, ...] = (
    "思想汇报总篇数",
    "思想汇报篇数",
    "思想汇报总数",
    "思想汇报总次数",
    "总篇数",
)

REPORT_COLUMN_ARABIC_PATTERN = re.compile(r"^\s*第\s*(\d+)\s*次\s*思想汇报\s*$")

NINE_PARTY_BRANCHES: dict[str, tuple[str, str]] = {
    "明理党支部": ("MINGLI", "明理党支部"),
    "德理党支部": ("DELI", "德理党支部"),
    "惟理党支部": ("WEILI", "惟理党支部"),
    "求理党支部": ("QIULI", "求理党支部"),
    "知理党支部": ("ZHILI", "知理党支部"),
    "昭理党支部": ("ZHAOLI", "昭理党支部"),
    "学理党支部": ("XUELI", "学理党支部"),
    "博理党支部": ("BOLI", "博理党支部"),
    "艺理党支部": ("YILI", "艺理党支部"),
}

STAGE_CHINESE_TO_CODE: dict[str, str] = {
    "入党积极分子": "ACTIVIST",
    "培养对象": "ACTIVIST",
    "中共预备党员": "PROBATIONARY",
    "预备党员": "PROBATIONARY",
    "正式党员": "FULL_MEMBER",
    "中共正式党员": "FULL_MEMBER",
}
VALID_STAGE_CODES: frozenset[str] = frozenset(("ACTIVIST", "PROBATIONARY", "FULL_MEMBER"))


@dataclass
class ReportColumn:
    sequence_number: int
    column_index: int
    source_column_name: str


@dataclass
class ColumnMapping:
    name_col: int | None = None
    student_number_col: int | None = None
    development_stage_col: int | None = None
    position_col: int | None = None
    applied_at_col: int | None = None
    reported_total_count_col: int | None = None
    report_columns: list[ReportColumn] = field(default_factory=list)

    def has_core_fields(self) -> bool:
        return (
            self.name_col is not None
            and self.student_number_col is not None
            and self.development_stage_col is not None
        )

    def missing_core_fields(self) -> list[str]:
        missing: list[str] = []
        if self.name_col is None:
            missing.append("姓名")
        if self.student_number_col is None:
            missing.append("学号")
        if self.development_stage_col is None:
            missing.append("发展阶段")
        return missing


@dataclass
class HeaderParseResult:
    mapping: ColumnMapping
    ok: bool
    error_code: str | None = None
    error_message: str | None = None
    unknown_columns: list[tuple[int, str]] = field(default_factory=list)


@dataclass
class RowParseResult:
    student_row: ParsedStudentRow | None
    errors: list[ParseError] = field(default_factory=list)
    warnings: list[ParseWarning] = field(default_factory=list)


@dataclass
class SheetParseResult:
    valid_rows: list[ParsedStudentRow] = field(default_factory=list)
    errors: list[ParseError] = field(default_factory=list)
    warnings: list[ParseWarning] = field(default_factory=list)


def _normalize(cell_value: object) -> str:
    if cell_value is None:
        return ""
    text = str(cell_value)
    return text.strip().replace("\u3000", " ")


def _cell_text(row: list[object], col_index: int | None) -> str:
    if col_index is None:
        return ""
    if col_index < 0 or col_index >= len(row):
        return ""
    return _normalize(row[col_index])


def _cell_raw(row: list[object], col_index: int | None) -> object:
    if col_index is None:
        return None
    if col_index < 0 or col_index >= len(row):
        return None
    return row[col_index]


def _find_col_index(
    header_cells: list[str],
    aliases: tuple[str, ...],
) -> int | None:
    for idx, cell in enumerate(header_cells):
        if cell in aliases:
            return idx
    return None


def _build_report_column_from_name(column_name: str, column_index: int) -> ReportColumn | None:
    parsed = parse_report_sequence(column_name)
    if parsed.ok and parsed.sequence_number is not None:
        return ReportColumn(
            sequence_number=parsed.sequence_number,
            column_index=column_index,
            source_column_name=parsed.source_column_name,
        )
    return None


def parse_header_row(header_row: list[object]) -> HeaderParseResult:
    cells = [_normalize(c) for c in header_row]

    name_col = _find_col_index(cells, NAME_ALIASES)
    student_number_col = _find_col_index(cells, STUDENT_NUMBER_ALIASES)
    development_stage_col = _find_col_index(cells, DEVELOPMENT_STAGE_ALIASES)
    position_col = _find_col_index(cells, POSITION_ALIASES)
    applied_at_col = _find_col_index(cells, APPLIED_AT_ALIASES)
    reported_total_count_col = _find_col_index(cells, REPORTED_TOTAL_COUNT_ALIASES)

    report_columns: list[ReportColumn] = []
    known_indices: set[int] = {
        i
        for i in (
            name_col,
            student_number_col,
            development_stage_col,
            position_col,
            applied_at_col,
            reported_total_count_col,
        )
        if i is not None
    }

    for idx, cell in enumerate(cells):
        if idx in known_indices or not cell:
            continue
        rc = _build_report_column_from_name(cell, idx)
        if rc is not None:
            report_columns.append(rc)

    report_columns.sort(key=lambda r: r.sequence_number)

    mapping = ColumnMapping(
        name_col=name_col,
        student_number_col=student_number_col,
        development_stage_col=development_stage_col,
        position_col=position_col,
        applied_at_col=applied_at_col,
        reported_total_count_col=reported_total_count_col,
        report_columns=report_columns,
    )

    if mapping.has_core_fields():
        return HeaderParseResult(mapping=mapping, ok=True)

    missing = mapping.missing_core_fields()
    messages: list[str] = []
    if "姓名" in missing:
        messages.append(_MSG_NAME_MISSING)
    if "学号" in missing:
        messages.append(_MSG_STUDENT_NUMBER_MISSING)
    if "发展阶段" in missing:
        messages.append(_MSG_STAGE_MISSING)
    error_message = "；".join(messages) if messages else "缺少核心表头字段"

    return HeaderParseResult(
        mapping=mapping,
        ok=False,
        error_code=ERROR_HEADER_NOT_FOUND,
        error_message=error_message,
    )


def normalize_development_stage(raw: object) -> tuple[str | None, str | None]:
    text = _normalize(raw)
    if not text:
        return None, _MSG_ROW_STAGE_EMPTY
    if text in VALID_STAGE_CODES:
        return text, None
    code = STAGE_CHINESE_TO_CODE.get(text)
    if code is not None:
        return code, None
    return None, _MSG_ROW_INVALID_STAGE


def _parse_int_or_none(raw: object) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw
    text = _normalize(raw)
    if not text:
        return None
    try:
        return int(text)
    except (TypeError, ValueError):
        return None


def _looks_like_chinese_name(text: str) -> bool:
    if not text:
        return False
    if not (2 <= len(text) <= 4):
        return False
    for ch in text:
        cp = ord(ch)
        if not (0x4E00 <= cp <= 0x9FFF):
            return False
    return True


def _looks_like_student_number(text: str) -> bool:
    if not text:
        return False
    if len(text) < 6 or len(text) > 12:
        return False
    return text.isdigit()


def _looks_like_date_text(text: str) -> bool:
    if not text:
        return False
    return bool(
        ("/" in text and len(text) >= 8)
        or ("-" in text and len(text) >= 8)
        or ("." in text and len(text) >= 8)
        or ("年" in text and "月" in text and "日" in text)
    )


def _looks_like_valid_stage_value(text: str) -> bool:
    if not text:
        return False
    if text in VALID_STAGE_CODES:
        return True
    if text in STAGE_CHINESE_TO_CODE:
        return True
    return False


def _cell_neighbour_text(row_values: list[object], col: int | None, delta: int) -> str:
    if col is None:
        return ""
    target = col + delta
    if target < 0 or target >= len(row_values):
        return ""
    return _normalize(row_values[target])


def _collect_shift_signals(
    row_values: list[object],
    mapping: ColumnMapping,
) -> list[str]:
    signals: list[str] = []

    name = _cell_text(row_values, mapping.name_col)
    if not name:
        right_name = _cell_neighbour_text(row_values, mapping.name_col, +1)
        if _looks_like_chinese_name(right_name):
            signals.append("姓名列为空，右邻列疑似姓名")

    student_number = _cell_text(row_values, mapping.student_number_col)
    if student_number and _looks_like_date_text(student_number):
        signals.append("学号列值疑似日期格式，可能列错位")
    if not student_number:
        right_number = _cell_neighbour_text(row_values, mapping.student_number_col, +1)
        if _looks_like_student_number(right_number):
            signals.append("学号列为空，右邻列疑似学号")

    if name and _looks_like_student_number(name):
        signals.append("姓名列值疑似学号（纯数字），可能列错位")

    stage_raw_text = _normalize(_cell_raw(row_values, mapping.development_stage_col))
    if stage_raw_text and not _looks_like_valid_stage_value(stage_raw_text):
        left_stage = _cell_neighbour_text(row_values, mapping.development_stage_col, -1)
        right_stage = _cell_neighbour_text(row_values, mapping.development_stage_col, +1)
        if _looks_like_valid_stage_value(left_stage) or _looks_like_valid_stage_value(right_stage):
            signals.append("发展阶段非法，但左右邻列出现合法阶段值")

    applied_text = _normalize(_cell_raw(row_values, mapping.applied_at_col))
    if mapping.applied_at_col is not None and applied_text:
        if _looks_like_chinese_name(applied_text) or _looks_like_student_number(applied_text):
            signals.append("申请入党时间列值类型不像日期")

    type_mismatches = 0
    if name and _looks_like_student_number(name):
        type_mismatches += 1
    if student_number and _looks_like_date_text(student_number):
        type_mismatches += 1
    if mapping.applied_at_col is not None and applied_text:
        if not _looks_like_date_text(applied_text) and not _looks_like_valid_stage_value(applied_text):
            name_or_num = _looks_like_chinese_name(applied_text) or _looks_like_student_number(applied_text)
            if name_or_num:
                type_mismatches += 1
    if stage_raw_text and not _looks_like_valid_stage_value(stage_raw_text):
        if _looks_like_date_text(stage_raw_text) or _looks_like_student_number(stage_raw_text):
            type_mismatches += 1
    if type_mismatches >= 2:
        signals.append(f"多个字段类型与表头明显不匹配（共{type_mismatches}处）")

    return signals


def detect_column_shift(
    row_values: list[object],
    mapping: ColumnMapping,
) -> tuple[bool, list[str]]:
    signals = _collect_shift_signals(row_values, mapping)
    return (len(signals) >= 2), signals


def parse_student_row(
    row_values: list[object],
    mapping: ColumnMapping,
    sheet_name: str,
    excel_row_number: int,
) -> RowParseResult:
    errors: list[ParseError] = []
    warnings: list[ParseWarning] = []

    shift_suspected, shift_signals = detect_column_shift(row_values, mapping)
    name_for_error = _cell_text(row_values, mapping.name_col)
    number_for_error = _cell_text(row_values, mapping.student_number_col)

    if shift_suspected:
        shift_detail = "；".join(shift_signals)
        shift_message = ERROR_MESSAGES.get(
            ERROR_ROW_COLUMN_SHIFT_SUSPECTED, "疑似列错位，本行跳过"
        )
        combined_message = f"{shift_message}（启发式信号：{shift_detail}）"
        errors.append(
            ParseError(
                code=ERROR_ROW_COLUMN_SHIFT_SUSPECTED,
                message=combined_message,
                sheet_name=sheet_name,
                excel_row_number=excel_row_number,
                student_name=name_for_error,
                student_number=number_for_error,
                field_name="row",
                source_value=shift_detail,
            )
        )
        return RowParseResult(student_row=None, errors=errors, warnings=warnings)

    name = name_for_error
    student_number = number_for_error
    stage_raw = _cell_raw(row_values, mapping.development_stage_col)
    position = _cell_text(row_values, mapping.position_col)
    applied_raw = _cell_raw(row_values, mapping.applied_at_col)
    reported_total_raw = _cell_raw(row_values, mapping.reported_total_count_col)

    def append_error(code: str, message: str, *, field_name: str = "", source_value: str = "") -> None:
        errors.append(
            ParseError(
                code=code,
                message=message,
                sheet_name=sheet_name,
                excel_row_number=excel_row_number,
                student_name=name,
                student_number=student_number,
                field_name=field_name,
                source_value=source_value,
            )
        )

    required_ok = True
    if not name:
        append_error(ERROR_ROW_MISSING_REQUIRED, _MSG_ROW_NAME_EMPTY, field_name="name")
        required_ok = False
    if not student_number:
        append_error(ERROR_ROW_MISSING_REQUIRED, _MSG_ROW_STUDENT_NUMBER_EMPTY, field_name="student_number")
        required_ok = False

    stage_code, stage_error = normalize_development_stage(stage_raw)
    if stage_error is not None:
        append_error(
            ERROR_ROW_INVALID_STAGE,
            stage_error,
            field_name="development_stage",
            source_value=_normalize(stage_raw),
        )
        required_ok = False

    if not required_ok:
        return RowParseResult(student_row=None, errors=errors, warnings=warnings)

    applied_at = None
    applied_raw_source = _normalize(applied_raw) if applied_raw is not None else ""
    if applied_raw_source:
        applied_result = parse_date(applied_raw)
        if not applied_result.ok:
            append_error(
                ERROR_ROW_INVALID_APPLIED_DATE,
                _MSG_ROW_INVALID_APPLIED_DATE,
                field_name="applied_at",
                source_value=applied_raw_source,
            )
            return RowParseResult(student_row=None, errors=errors, warnings=warnings)
        applied_at = applied_result.value

    reported_total_count = _parse_int_or_none(reported_total_raw)
    reported_total_raw_text = _normalize(reported_total_raw) if reported_total_raw is not None else ""
    if reported_total_raw_text and reported_total_count is None:
        append_error(
            error_codes.ERROR_REPORT_TOTAL_INVALID,
            f"思想汇报总篇数非法：{reported_total_raw_text}",
            field_name="reported_total_count",
            source_value=reported_total_raw_text,
        )
        return RowParseResult(student_row=None, errors=errors, warnings=warnings)

    report_items: list[ParsedReportItem] = []
    for rc in mapping.report_columns:
        cell_raw = _cell_raw(row_values, rc.column_index)
        cell_text = _normalize(cell_raw)
        if not cell_text:
            continue
        date_result = parse_date(cell_raw)
        if date_result.ok and date_result.value is not None:
            report_items.append(
                ParsedReportItem(
                    sequence_number=rc.sequence_number,
                    submitted_at=date_result.value,
                    source_column_name=rc.source_column_name,
                )
            )
        else:
            warnings.append(
                ParseWarning(
                    code=WARNING_REPORT_DATE_INVALID,
                    message=_MSG_REPORT_DATE_INVALID_TEMPLATE.format(
                        source_column=rc.source_column_name,
                    ),
                    sheet_name=sheet_name,
                    excel_row_number=excel_row_number,
                    student_name=name,
                    student_number=student_number,
                    field_name=f"report:{rc.source_column_name}",
                    source_value=cell_text,
                    parsed_value="",
                )
            )

    calculated_date_count = len(report_items)

    if reported_total_count is not None and calculated_date_count != reported_total_count:
        warnings.append(
            ParseWarning(
                code=WARNING_REPORT_COUNT_MISMATCH,
                message=_MSG_REPORT_COUNT_MISMATCH_TEMPLATE.format(
                    reported=reported_total_count,
                    calculated=calculated_date_count,
                ),
                sheet_name=sheet_name,
                excel_row_number=excel_row_number,
                student_name=name,
                student_number=student_number,
                field_name="reported_total_count",
                source_value=str(reported_total_count),
                parsed_value=str(calculated_date_count),
            )
        )

    student_row = ParsedStudentRow(
        sheet_name=sheet_name,
        excel_row_number=excel_row_number,
        branch_code="",
        branch_name="",
        name=name,
        student_number=student_number,
        development_stage=stage_code or "",
        position=position,
        applied_at=applied_at,
        reported_total_count=reported_total_count,
        calculated_date_count=calculated_date_count,
        report_items=report_items,
        warnings=warnings[:],
    )

    return RowParseResult(student_row=student_row, errors=errors, warnings=warnings)


def parse_sheet_rows(
    data_rows: list[list[object]],
    mapping: ColumnMapping,
    sheet_name: str,
    *,
    start_excel_row_number: int = 3,
) -> SheetParseResult:
    sheet = SheetParseResult()
    for offset, row in enumerate(data_rows):
        excel_row_number = start_excel_row_number + offset
        outcome = parse_student_row(
            row_values=row,
            mapping=mapping,
            sheet_name=sheet_name,
            excel_row_number=excel_row_number,
        )
        sheet.errors.extend(outcome.errors)
        sheet.warnings.extend(outcome.warnings)
        if outcome.student_row is not None:
            sheet.valid_rows.append(outcome.student_row)
    return sheet

def parse_workbook(file_path: Path) -> ParseResult:
    """唯一生产入口：打开 Excel 工作簿，遍历全部工作表，聚合解析结果。

    输入：本地 Excel 文件 Path。
    输出：ParseResult。
    异常：文件不存在抛 FileNotFoundError；IO/损坏/openpyxl 不可读属系统异常。
    """
    if not file_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{file_path}")

    wb = load_workbook(file_path, data_only=True, read_only=False)
    try:
        result = ParseResult()
        sheet_names = wb.sheetnames
        result.total_sheets = len(sheet_names)

        for title in sheet_names:
            ws = wb[title]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]

            branch_info = NINE_PARTY_BRANCHES.get(title)

            if branch_info is None:
                result.errors.append(
                    ParseError(
                        code=ERROR_UNKNOWN_SHEET,
                        message=f"工作表名称不在九个党支部映射中：{title}",
                        sheet_name=title,
                        excel_row_number=1,
                        student_name="",
                        student_number="",
                        field_name="sheet",
                        source_value=title,
                    )
                )
                result.sheet_results.append(
                    SheetResult(
                        sheet_name=title,
                        branch_code=None,
                        branch_name=title,
                        status="unknown",
                        total_rows=0,
                        valid_row_count=0,
                        error_count=1,
                        warning_count=0,
                    )
                )
                continue

            branch_code, branch_name = branch_info

            if len(rows) < 2:
                result.errors.append(
                    ParseError(
                        code=ERROR_HEADER_NOT_FOUND,
                        message="工作表行数不足，无法识别表头",
                        sheet_name=title,
                        excel_row_number=1,
                        student_name="",
                        student_number="",
                        field_name="header",
                        source_value="",
                    )
                )
                result.sheet_results.append(
                    SheetResult(
                        sheet_name=title,
                        branch_code=branch_code,
                        branch_name=branch_name,
                        status="failed",
                        total_rows=0,
                        valid_row_count=0,
                        error_count=1,
                        warning_count=0,
                    )
                )
                continue

            header_result = parse_header_row(rows[1])
            if not header_result.ok:
                result.errors.append(
                    ParseError(
                        code=header_result.error_code or ERROR_HEADER_NOT_FOUND,
                        message=header_result.error_message or ERROR_MESSAGES[ERROR_HEADER_NOT_FOUND],
                        sheet_name=title,
                        excel_row_number=2,
                        student_name="",
                        student_number="",
                        field_name="header",
                        source_value=" | ".join(str(c) for c in rows[1]),
                    )
                )
                result.sheet_results.append(
                    SheetResult(
                        sheet_name=title,
                        branch_code=branch_code,
                        branch_name=branch_name,
                        status="failed",
                        total_rows=0,
                        valid_row_count=0,
                        error_count=1,
                        warning_count=0,
                    )
                )
                continue

            mapping = header_result.mapping
            data_rows = rows[2:]
            sheet_parse = parse_sheet_rows(
                data_rows=data_rows,
                mapping=mapping,
                sheet_name=title,
                start_excel_row_number=3,
            )

            for row in sheet_parse.valid_rows:
                row.branch_code = branch_code
                row.branch_name = branch_name

            if mapping.reported_total_count_col is None:
                sheet_parse.warnings.append(
                    ParseWarning(
                        code=WARNING_REPORT_TOTAL_COLUMN_MISSING,
                        message=_MSG_REPORT_TOTAL_COLUMN_MISSING,
                        sheet_name=title,
                        excel_row_number=2,
                        student_name="",
                        student_number="",
                        field_name="reported_total_count",
                        source_value="",
                        parsed_value="",
                    )
                )

            non_empty_data_rows = sum(1 for r in data_rows if any(c is not None and str(c).strip() != "" for c in r))
            row_level_errors = len(sheet_parse.errors)
            sheet_warnings = len(sheet_parse.warnings)
            sheet_status = "success" if len(sheet_parse.valid_rows) > 0 or non_empty_data_rows == 0 else "failed"

            result.sheet_results.append(
                SheetResult(
                    sheet_name=title,
                    branch_code=branch_code,
                    branch_name=branch_name,
                    status=sheet_status,
                    total_rows=non_empty_data_rows,
                    valid_row_count=len(sheet_parse.valid_rows),
                    error_count=row_level_errors,
                    warning_count=sheet_warnings,
                )
            )
            result.valid_rows.extend(sheet_parse.valid_rows)
            result.errors.extend(sheet_parse.errors)
            result.warnings.extend(sheet_parse.warnings)

        result.total_rows = sum(s.total_rows for s in result.sheet_results if s.status != "unknown")
        result.success_rows = len(result.valid_rows)
        row_level_error_codes = {
            ERROR_ROW_MISSING_REQUIRED,
            ERROR_ROW_INVALID_STAGE,
            ERROR_ROW_INVALID_APPLIED_DATE,
            ERROR_ROW_COLUMN_SHIFT_SUSPECTED,
        }
        result.skipped_rows = sum(1 for e in result.errors if e.code in row_level_error_codes)
        result.warning_rows = len({
            (w.sheet_name, w.excel_row_number) for w in result.warnings
        })
        result.success_sheets = sum(1 for s in result.sheet_results if s.status == "success")
        result.failed_sheets = sum(1 for s in result.sheet_results if s.status == "failed")

        return result
    finally:
        wb.close()
