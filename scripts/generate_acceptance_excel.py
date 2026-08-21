from __future__ import annotations

import argparse
import json
import os
import random
import zipfile
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook


BRANCHES = (
    ("MINGLI", "明理党支部"),
    ("DELI", "德理党支部"),
    ("WEILI", "惟理党支部"),
    ("QIULI", "求理党支部"),
    ("ZHILI", "知理党支部"),
    ("ZHAOLI", "昭理党支部"),
    ("XUELI", "学理党支部"),
    ("BOLI", "博理党支部"),
    ("YILI", "艺理党支部"),
)
STAGES = ("入党积极分子", "中共预备党员", "正式党员")
HEADERS = (
    "姓名",
    "学号",
    "发展阶段",
    "职务",
    "申请入党时间",
    "思想汇报总篇数",
    "第一次思想汇报",
    "第二次思想汇报",
    "第三次思想汇报",
)


@dataclass(frozen=True)
class AcceptanceDataset:
    seed: int
    student_count: int
    warning_rows: int
    first_workbook: str
    second_workbook: str
    invalid_workbook: str
    duplicate_workbook: str


def generate_acceptance_dataset(
    output_directory: Path,
    *,
    student_count: int = 1500,
    seed: int = 20260822,
) -> AcceptanceDataset:
    if student_count < len(BRANCHES):
        raise ValueError("合成学生数量不得少于九个党支部数量。")
    output_directory.mkdir(parents=True, exist_ok=True)
    randomizer = random.Random(seed)
    records = _student_records(student_count, randomizer)
    first = output_directory / "acceptance_first.xlsx"
    second = output_directory / "acceptance_second.xlsx"
    invalid = output_directory / "acceptance_invalid.xlsx"
    duplicate = output_directory / "acceptance_duplicate.xlsx"
    _write_main_workbook(first, records, second_import=False)
    _write_main_workbook(second, records, second_import=True)
    _write_invalid_workbook(invalid)
    _write_duplicate_workbook(duplicate, records[0])
    dataset = AcceptanceDataset(
        seed=seed,
        student_count=student_count,
        warning_rows=sum(1 for record in records if record["warning"]),
        first_workbook=str(first),
        second_workbook=str(second),
        invalid_workbook=str(invalid),
        duplicate_workbook=str(duplicate),
    )
    (output_directory / "dataset_manifest.json").write_text(
        json.dumps(asdict(dataset), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return dataset


def _student_records(student_count: int, randomizer: random.Random) -> list[dict]:
    records: list[dict] = []
    base_date = date(2022, 1, 1)
    for index in range(student_count):
        branch_code, branch_name = BRANCHES[index % len(BRANCHES)]
        first_report = base_date + timedelta(days=index % 300)
        records.append(
            {
                "branch_code": branch_code,
                "branch_name": branch_name,
                "name": f"合成学生{index + 1:04d}",
                "student_number": f"SYN{index + 1:07d}",
                "stage": STAGES[index % len(STAGES)],
                "position": f"测试职务{randomizer.randint(1, 12):02d}",
                "applied_at": base_date + timedelta(days=index % 365),
                "reported_total": 4 if index % 100 == 0 else 3,
                "reports": (
                    first_report,
                    first_report + timedelta(days=90),
                    first_report + timedelta(days=180),
                ),
                "warning": index % 100 == 0,
            }
        )
    return records


def _new_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    fixed_timestamp = datetime(2026, 8, 22, 0, 0, 0)
    workbook.properties.created = fixed_timestamp
    workbook.properties.modified = fixed_timestamp
    workbook.properties.creator = "party-building-archive-system acceptance generator"
    return workbook


def _add_header(sheet) -> None:
    sheet.append(["学生信息", "学生信息", "发展信息", "发展信息", "材料", "材料", "思想汇报", "思想汇报", "思想汇报"])
    sheet.append(list(HEADERS))


def _write_main_workbook(path: Path, records: list[dict], *, second_import: bool) -> None:
    workbook = _new_workbook()
    sheets = {}
    for _, branch_name in BRANCHES:
        sheet = workbook.create_sheet(branch_name)
        _add_header(sheet)
        sheets[branch_name] = sheet
    for record in records:
        reports = record["reports"]
        if second_import:
            reports = tuple(item + timedelta(days=30) for item in reports)
        sheets[record["branch_name"]].append(
            [
                record["name"],
                record["student_number"],
                record["stage"],
                "第二次导入职务" if second_import else record["position"],
                record["applied_at"],
                record["reported_total"],
                *reports,
            ]
        )
    workbook.save(path)
    workbook.close()
    _normalize_xlsx_archive(path)


def _write_invalid_workbook(path: Path) -> None:
    workbook = _new_workbook()
    sheet = workbook.create_sheet("明理党支部")
    _add_header(sheet)
    sheet.append(["非法阶段", "NEG0001", "未知阶段", "", "2024/01/01", 0, None, None, None])
    sheet.append(["非法日期", "NEG0002", "入党积极分子", "", "2024/99/99", 0, None, None, None])
    sheet.append([None, "NEG0003", "入党积极分子", "", "2024/01/01", 0, None, None, None])
    empty = workbook.create_sheet("德理党支部")
    _add_header(empty)
    unknown = workbook.create_sheet("未知支部")
    _add_header(unknown)
    unknown.append(["未知支部学生", "NEG0004", "入党积极分子", "", "2024/01/01", 0, None, None, None])
    failed = workbook.create_sheet("惟理党支部")
    failed.append(["错误分组"])
    failed.append(["非标准表头"])
    failed.append(["无法解析的数据"])
    workbook.save(path)
    workbook.close()
    _normalize_xlsx_archive(path)


def _write_duplicate_workbook(path: Path, record: dict) -> None:
    workbook = _new_workbook()
    sheet = workbook.create_sheet(record["branch_name"])
    _add_header(sheet)
    base = [
        record["name"],
        record["student_number"],
        record["stage"],
        record["position"],
        record["applied_at"],
        record["reported_total"],
        *record["reports"],
    ]
    sheet.append(base)
    sheet.append(["重复学号学生", *base[1:]])
    workbook.save(path)
    workbook.close()
    _normalize_xlsx_archive(path)


def _normalize_xlsx_archive(path: Path) -> None:
    """固定ZIP元数据，使相同种子在Windows/Linux生成字节一致的XLSX。"""
    temporary = path.with_name(f".{path.name}.normalized.tmp")
    with zipfile.ZipFile(path, "r") as source:
        entries = [(name, source.read(name)) for name in sorted(source.namelist())]
    try:
        with zipfile.ZipFile(
            temporary,
            "w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=9,
        ) as destination:
            for name, payload in entries:
                info = zipfile.ZipInfo(name, date_time=(2026, 8, 22, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = 0
                info.external_attr = 0o600 << 16
                destination.writestr(info, payload, compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)
        os.replace(temporary, path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise


def main() -> None:
    parser = argparse.ArgumentParser(description="生成不含真实个人信息的Excel联合验收数据。")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--student-count", type=int, default=1500)
    parser.add_argument("--seed", type=int, default=20260822)
    args = parser.parse_args()
    dataset = generate_acceptance_dataset(
        args.output_dir.resolve(),
        student_count=args.student_count,
        seed=args.seed,
    )
    print(json.dumps(asdict(dataset), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
